#import config module for environmental variability
import config
#import my utility class and function
import MyUtility
# Import dictionary from file COG_name.py
from COG_name import get_dictionary
# Get dictionary
COG_dict = get_dictionary()

#import threading
from threading import Thread
#pandas import
import pandas as pd
#numpy import
import numpy as np

#tkinter import
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter.filedialog import askopenfile
from tkinter.messagebox import showinfo
from tkinter.ttk import Separator, Style

# Standard library packages
import io
import os
import sys
import openpyxl
import csv

# Import Biopython modules to interact with KEGG
#from Bio import SeqIO
#from Bio.KEGG import REST
#from Bio.KEGG.KGML import KGML_parser
#from Bio.Graphics.KGML_vis import KGMLCanvas

#Import module to request
from urllib.request import urlopen
import ssl

#utility for load image
def resource_path(relative_path):
  try:
    base_path = sys._MEIPASS
  except Exception:
    base_path = os.path.abspath(".")

  return os.path.join(base_path, relative_path)

def resource_path_2(relative_path):
  base_path = getattr(
    sys,
    '_MEIPASS',
    os.path.dirname(os.path.abspath(__file__)) )
  
  return os.path.join(base_path, relative_path)

#Request for KEGG data
def _q(op, arg1, arg2=None, arg3=None):
  URL = "https://rest.kegg.jp/%s"
  if arg2 and arg3:
    args = f"{op}/{arg1}/{arg2}/{arg3}"
  elif arg2:
    args = f"{op}/{arg1}/{arg2}"
  else:
    args = f"{op}/{arg1}"

  #original request
  #resp = urlopen(URL % (args))
  
  #edit request to avoid certificate request      
  ctx = ssl.create_default_context()
  ctx.check_hostname = False
  ctx.verify_mode = ssl.CERT_NONE
  resp=urlopen(URL % (args), context=ctx)

  if "image" == arg2:
      return resp

  handle = io.TextIOWrapper(resp, encoding="UTF-8")
  handle.url = resp.url
  return handle

#query for KEGG list   
def kegg_list(database, org=None):
  """KEGG list - Entry list for database, or specified database entries.
  db - database or organism (string)
  org - optional organism (string), see below.
  For the pathway and module databases the optional organism can be
  used to restrict the results.
  """
  # TODO - split into two functions (dbentries seems separate)?
  #
  #  https://rest.kegg.jp/list/<database>/<org>
  #
  #  <database> = pathway | module
  #  <org> = KEGG organism code
  if database in ("pathway", "module") and org:
    resp = _q("list", database, org)
  elif isinstance(database, str) and database and org:
    raise ValueError("Invalid database arg for kegg list request.")

  # https://rest.kegg.jp/list/<database>
  #
  # <database> = pathway | brite | module | disease | drug | environ |
  #              ko | genome | <org> | compound | glycan | reaction |
  #              rpair | rclass | enzyme | organism
  # <org> = KEGG organism code or T number
  #
  #
  # https://rest.kegg.jp/list/<dbentries>
  #
  # <dbentries> = KEGG database entries involving the following <database>
  # <database> = pathway | brite | module | disease | drug | environ |
  #              ko | genome | <org> | compound | glycan | reaction |
  #              rpair | rclass | enzyme
  # <org> = KEGG organism code or T number
  else:
    if isinstance(database, list):
      if len(database) > 100:
        raise ValueError(
          "Maximum number of databases is 100 for kegg list query"
        )
      database = ("+").join(database)
    resp = _q("list", database)

  return resp

#Some code to return a Pandas dataframe, given tabular text
def to_df(result):
  #
  return pd.read_table(io.StringIO(result), header=None)

#Manage KEGG data
def manage_kegg_query(self):
  if( hasattr(self, 'query_ko') ): #category_to_search['KEGG_ko']):
    #preapre df
    self.df_ko = pd.DataFrame()
    #get online information from kegg.com and convert into dataframe
    self.df_ko = to_df(self.query_ko)
    #take information about ko codes from originale df
    need_df = (self.df.assign(KEGG_ko=self.df['KEGG_ko'].str.split('[,;]')).explode('KEGG_ko'))
    tmp_list = need_df['KEGG_ko'].dropna().unique().tolist()
    #delete everything is not on the list
    self.df_ko = self.df_ko[self.df_ko[0].isin(tmp_list)]
    #edit ko name to remove unuseless information
    # Rimuovi il carattere ";" e tutto quello che lo precede
    self.df_ko[1] = self.df_ko[1].str.replace('.*; ', '', regex=True)
    # Rimuovi il testo "[EC" e quello che lo segue
    self.df_ko[1] = self.df_ko[1].str.replace(' \[EC.*\]', '', regex=True)
    
  if( hasattr(self, 'query_pathway') ): #category_to_search['KEGG_Pathway']):
    #preapre df
    self.df_pathway = pd.DataFrame()
    #get online information from kegg.com and convert into dataframe
    self.df_pathway = to_df(self.query_pathway)
    #take information about pathway codes from originale df
    need_df = (self.df.assign(KEGG_Pathway=self.df['KEGG_Pathway'].str.split('[,;]')).explode('KEGG_Pathway'))
    tmp_list = need_df['KEGG_Pathway'].dropna().unique().tolist()
    #delete everything is not on the list
    self.df_pathway = self.df_pathway[self.df_pathway[0].isin(tmp_list)]

  if( hasattr(self, 'query_module') ): #category_to_search['KEGG_Module']):
    #preapre df
    self.df_module = pd.DataFrame()
    #get online information from kegg.com and convert into dataframe
    self.df_module = to_df(self.query_module)
    #take information about module codes from originale df
    need_df = (self.df.assign(KEGG_Module=self.df['KEGG_Module'].str.split('[,;]')).explode('KEGG_Module'))
    tmp_list = need_df['KEGG_Module'].dropna().unique().tolist()
    #delete everything is not on the list
    self.df_module = self.df_module[self.df_module[0].isin(tmp_list)]

  if( hasattr(self, 'query_reaction') ): #category_to_search['KEGG_Reaction']):
    #preapre df
    self.df_reaction = pd.DataFrame()
    #get online information from kegg.com and convert into dataframe
    self.df_reaction = to_df(self.query_reaction)
    #take information about reactions codes from originale df
    need_df = (self.df.assign(KEGG_Reaction=self.df['KEGG_Reaction'].str.split('[,;]')).explode('KEGG_Reaction'))
    tmp_list = need_df['KEGG_Reaction'].dropna().unique().tolist()
    #delete everything is not on the list
    self.df_reaction = self.df_reaction[self.df_reaction[0].isin(tmp_list)]

#class to upload file
class AsyncUpload(Thread):
  def __init__(self, filepath):
    super().__init__()

    self.filepath = filepath

  def run(self):
    #variable to check if file will be open
    self.fileOpen = True
    #open file with pandas
    try:
      #save file with pandas
      file_extension = self.filepath.split(".")[-1]
      if file_extension == "xlsx":
        self.df = pd.read_excel(self.filepath)
      else:
        self.df = pd.read_csv(self.filepath, sep='\t', low_memory=False)
    except Exception as e:
      #print("===>>" + str(e))
      self.fileOpen = False

#class to upload file mzTab
class AsyncUpload_mzTab(Thread):
  def __init__(self, filepath, headerName, row_name):
    super().__init__()

    self.filepath = filepath
    self.headerName = headerName
    self.row_name = row_name

  def run(self):
    #variable to check if file will be open
    self.fileOpen = True

    try:
      # Apri il file .mzTab
      with open(self.filepath, 'r') as f:
          # Inizializza le liste per le righe headerName e row_name
          ppp_rows = []
          mmm_rows = []

          # Scorri il file riga per riga
          for line in f:
              # Estra le righe che iniziano con headerName e row_name
              if line.startswith(self.headerName):
                  ppp_rows.append(line.strip().split('\t'))
              elif line.startswith(self.row_name):
                  mmm_rows.append(line.strip().split('\t'))

      #verifico se posso creare il dataframe
      if(len(ppp_rows) == 0):
        self.badFile = True
      else:
        # Crea il dataframe con i nomi delle colonne
        self.df = pd.DataFrame(mmm_rows, columns=ppp_rows[0])

    except Exception as e:
      #print("===>>" + str(e))
      self.fileOpen = False

#class to upload file exluding some start and end line
class AsyncUpload_2(Thread):
  def __init__(self, filepath):
    super().__init__()

    self.filepath = filepath

  def run(self):
    #variable to check if file will be open
    self.fileOpen = True

    #open file with pandas
    try:
      #save file with pandas
      file_extension = self.filepath.split(".")[-1]
      if file_extension == "xlsx":
        #open file
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        #count initial and final '#'
        count_start = 0
        count_end = 0
        #count initial '#'
        for row in ws.iter_rows():
          if row[0].value is not None and str(row[0].value).startswith("#"):
            count_start += 1
          else:
            break
        #count final '#'    
        for row in reversed(list(ws.iter_rows())):
          if row[0].value is not None and str(row[0].value).startswith("#"):
            count_end += 1
          else:
            break

        #read removing first and last row that start with '#' because are not used
        self.df = pd.read_excel(self.filepath, skiprows=count_start, skipfooter = count_end)

      else:
        # apri il file csv in lettura e crea un reader csv
        with open(self.filepath, 'r') as csvfile:
            reader = csv.reader(csvfile, delimiter='\t')

            # inizializza la variabile per l'intestazione
            header = None
            found_header = False

            # crea una lista contenente tutte le righe che non iniziano per '#' (tranne '#query' se è l'intestazione)
            rows = []
            for row in reader:
                if row[0].startswith('#query') and not found_header:
                    header = row
                    found_header = True
                elif not row[0].startswith('#'):
                    rows.append(row)

        # crea un dataframe pandas dal file csv, usando l'intestazione corretta (se trovata) o la prima riga dei dati
        if header is not None:
            self.df = pd.DataFrame(rows, columns=header)

            # Estrae il nome completo delle colonne che iniziano con '#query'
            query_columns = self.df.filter(regex='#query').columns
            query_columns_names = [col for col in query_columns]
            #rename the column
            if(len(query_columns_names) == 1):
              self.df = self.df.rename(columns={query_columns_names[0]: 'query'})
        else:
            self.df = pd.DataFrame(rows[1:], columns=rows[0])

    except Exception as e:
      #print("===>>" + str(e))
      self.fileOpen = False

#class to download file
class AsyncDownload(Thread):
  def __init__(self, df_tmp, file_path):
    super().__init__()

    self.df_tmp = df_tmp
    self.file_path = file_path

  def run(self):
    #variable to check if file will be saved
    self.fileSaved = True
    #open file with pandas
    try:
      #save file with pandas
      file_extension = self.file_path.split(".")[-1]
      #file_path_without_extension, file_extension = self.file.name.rsplit(".", 1)
      if file_extension == "xlsx":
        self.df_tmp.to_excel(self.file_path, index=False)
      else:
        self.df_tmp.to_csv(self.file_path, sep='\t', index=False, header=True,)
           
    except Exception as e:
      #print("===>>" + str(e))
      self.fileSaved = False

#class to downalod Summary Metrics file for input
class ManageSummaryMetricsPre(Thread):
  def __init__(self, window):
    super().__init__()

    #take window data
    self.window = window

  def run(self):
    #take a copy of window to do controls
    window = self.window

    #get abundance colums name
    abundance_set = list(window.df.filter(regex=r'F\d+'))

    #Remove all rows that have unassigned in all abundance
    condizione_colonne = window.df[abundance_set] == 'unassigned'
    # Verifica se tutte le colonne soddisfano la condizione (tutte True lungo l'asse 1)
    condizione_generale = condizione_colonne.all(axis=1)
    # Seleziona solo le righe che soddisfano la condizione generale
    df_filtrato = window.df.loc[~condizione_generale]

    # Convertire le colonne in numerico
    window.df[abundance_set] = window.df[abundance_set].apply(pd.to_numeric, errors='coerce')

    #Remove all rows that have NaN in all abundance
    condizione_colonne = window.df[abundance_set].isna()  # Utilizza .isna() o .isnull()
    # Verifica se tutte le colonne soddisfano la condizione (tutte True lungo l'asse 1)
    condizione_generale = condizione_colonne.all(axis=1)
    # Seleziona solo le righe che soddisfano la condizione generale
    df_filtrato = window.df.loc[~condizione_generale]

    #create new df with the name of aboundances
    new_df = pd.DataFrame(columns=["Metrics"] + abundance_set + ["Whole dataset"])

    # Creare una lista vuota per contenere i DataFrame da concatenare
    dfs_to_concat_count = []
    dfs_to_concat_sum = []

    ##get correct input type
    column_count_text = ""
    column_total_text = ""
    if(MyUtility.workDict["mode"] == "Proteins"):
      column_count_text = "Quantified proteins"
      column_total_text = "Total abundance"
    elif(MyUtility.workDict["mode"] == "Peptides"):
        column_count_text = "Quantified peptides"
        column_total_text = "Total abundance"
    else:
        column_count_text = "Identified peptides"
        column_total_text = "Total PSMs"

    ##### Quantified proteins #####
    whole_count_tot = len(window.df)
    ### Count ###
    # Calcolo il numero di valori > 0 e diversi da NaN per ogni colonna 'val_x'
    count_vals = window.df[abundance_set].gt(0).sum()
    # Creo un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
    new_row = {'Metrics': column_count_text}
    new_row.update(count_vals.to_dict())
    new_row.update({'Whole dataset': whole_count_tot})
    # Aggiungo la nuova riga al DataFrame 'new_df'
    tmp_df = pd.DataFrame(new_row, index=[0])
    #aggiungo al vettore dei risultati
    dfs_to_concat_count.append(tmp_df)
    ### Sum ###
    # Calcolo la somma di valori > 0 e diversi da NaN per ogni colonna 'val_x'
    count_vals = window.df[abundance_set].sum()
    # Creo un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
    new_row = {'Metrics': column_total_text}
    new_row.update(count_vals.to_dict())
    #new_row.update({'Whole dataset': whole_count_tot})
    # Aggiungo la nuova riga al DataFrame 'new_df'
    tmp_df = pd.DataFrame(new_row, index=[0])
    #aggiungo al vettore dei risultati
    dfs_to_concat_sum.append(tmp_df)
    
    ##### Marked as #####
    if 'Marked as' in window.df.columns:
      # Ottenere un array degli elementi unici nella colonna 'maked as'
      window.df['Marked as'] = window.df['Marked as'].astype(str)
      unique_markedas = sorted(window.df['Marked as'].unique())
      # Iterare sugli elementi unici
      for element in unique_markedas:
        # Filtrare il DataFrame per includere solo le righe in cui 'Marked as' è uguale a 'element' e non ci sono spazi vuoti
        filtered_df = window.df[(window.df['Marked as'] == element) & (window.df['Marked as'] != '') & (window.df['Marked as'].notna()) & (window.df['Marked as'] != 'unassigned')]

        #conto il totale delle righe che contengono il valore del quale conto le metriche
        whole_count = filtered_df['Marked as'].count()

        ### Count ###
        # Calcolare il numero di valori > 0 e diversi da NaN per ogni colonna 'val_x' solo nelle righe filtrate
        count_vals = filtered_df[abundance_set].gt(0).sum()

        # Creare un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
        new_row = {'Metrics': column_count_text+' - ' + element}
        new_row.update(count_vals.to_dict())
        new_row.update({'Whole dataset': whole_count})

        # Creare un DataFrame con la riga corrente
        tmp_df = pd.DataFrame(new_row, index=[0])

        # Aggiungere il DataFrame corrente alla lista di DataFrame da concatenare
        dfs_to_concat_count.append(tmp_df)

        ### Sum ###
        # Calcolare il numero di valori > 0 e diversi da NaN per ogni colonna 'val_x' solo nelle righe filtrate
        count_vals = filtered_df[abundance_set].sum()

        # Creare un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
        new_row = {'Metrics': column_total_text+' - ' + element}
        new_row.update(count_vals.to_dict())
        #new_row.update({'Whole dataset': whole_count})

        # Creare un DataFrame con la riga corrente
        tmp_df = pd.DataFrame(new_row, index=[0])

        # Aggiungere il DataFrame corrente alla lista di DataFrame da concatenare
        dfs_to_concat_sum.append(tmp_df)

    ##### Count taxonomic count #####
    if 'taxonomic_table' in MyUtility.workDict:
      for column in MyUtility.workDict['taxonomic_table']:
        # Filtrare il DataFrame per includere solo le righe in cui nella colonna selezionata è presente un valore
        filtered_df = window.df[(window.df[column] != '') & (window.df[column].notna()) & (window.df[column] != 'unassigned')]

        #conto il totale delle righe che contengono il valore del quale conto le metriche
        whole_count = filtered_df[column].count()

        ### Count ###
        # Calcolare il numero di valori > 0 e diversi da NaN per ogni colonna 'val_x' solo nelle righe filtrate
        count_vals = filtered_df[abundance_set].gt(0).sum()

        # Creare un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
        new_row = {'Metrics': column_count_text+' - ' + column}
        new_row.update(count_vals.to_dict())
        new_row.update({'Whole dataset': whole_count})

        # Creare un DataFrame con la riga corrente
        tmp_df = pd.DataFrame(new_row, index=[0])

        # Aggiungere il DataFrame corrente alla lista di DataFrame da concatenare
        dfs_to_concat_count.append(tmp_df)

        ### Sum ###
        # Calcolare il numero di valori > 0 e diversi da NaN per ogni colonna 'val_x' solo nelle righe filtrate
        count_vals = filtered_df[abundance_set].sum()

        # Creare un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
        new_row = {'Metrics': column_total_text+' - ' + column}
        new_row.update(count_vals.to_dict())
        #new_row.update({'Whole dataset': whole_count})

        # Creare un DataFrame con la riga corrente
        tmp_df = pd.DataFrame(new_row, index=[0])

        # Aggiungere il DataFrame corrente alla lista di DataFrame da concatenare
        dfs_to_concat_sum.append(tmp_df)

    ##### Count functional count #####
    if 'functional_table' in MyUtility.workDict:
      for column in MyUtility.workDict['functional_table']:
        # Filtrare il DataFrame per includere solo le righe in cui nella colonna selezionata è presente un valore
        filtered_df = window.df[(window.df[column] != '') & (window.df[column].notna()) & (window.df[column] != 'unassigned')]

        #conto il totale delle righe che contengono il valore del quale conto le metriche
        whole_count = filtered_df[column].count()
        
        ### Count ###
        # Calcolare il numero di valori > 0 e diversi da NaN per ogni colonna 'val_x' solo nelle righe filtrate
        count_vals = filtered_df[abundance_set].gt(0).sum()

        # Creare un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
        new_row = {'Metrics': column_count_text+' - ' + column}
        new_row.update(count_vals.to_dict())
        new_row.update({'Whole dataset': whole_count})

        # Creare un DataFrame con la riga corrente
        tmp_df = pd.DataFrame(new_row, index=[0])

        # Aggiungere il DataFrame corrente alla lista di DataFrame da concatenare
        dfs_to_concat_count.append(tmp_df)

        ### Sum ###
        # Calcolare il numero di valori > 0 e diversi da NaN per ogni colonna 'val_x' solo nelle righe filtrate
        count_vals = filtered_df[abundance_set].sum()

        # Creare un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
        new_row = {'Metrics': column_total_text+' - ' + column}
        new_row.update(count_vals.to_dict())
        #new_row.update({'Whole dataset': whole_count})

        # Creare un DataFrame con la riga corrente
        tmp_df = pd.DataFrame(new_row, index=[0])

        # Aggiungere il DataFrame corrente alla lista di DataFrame da concatenare
        dfs_to_concat_sum.append(tmp_df)
    
    ##### Add all row in new_df #####
    # Concatenare tutti i DataFrame nella lista in un unico DataFrame
    new_df = pd.concat(dfs_to_concat_count+dfs_to_concat_sum, ignore_index=True)

    #save df tmp in the window
    window.df_tmp = new_df

#class to download aggregation files
class AsyncDownload_Aggregation(Thread):
  def __init__(self, window, df, my_list, params, file_path):
    super().__init__()

    #take a copy of window to do controls
    self.window = window 
    #save the df recived
    self.df = df
    #save the list recived
    self.my_list = my_list
    #params to work
    self.params = params
    #save file_path
    self.file_path = file_path
    #variable to know if user want replace all existing file with the same path
    #2=to ask;  1=positive answer;  0=negative answer
    self.replaceAll = 2;

  def run(self):
    ### only for Summary metrics ###
    # Creare una lista vuota per contenere i DataFrame da concatenare
    dfs_to_concat = []
    #get abundance colums name
    abundance_set = list(self.df.filter(regex=r'F\d+'))
    # Convertire le colonne in numerico
    self.df[abundance_set] = self.df[abundance_set].apply(pd.to_numeric, errors='coerce')
    #create new df with the name of aboundances
    metrics_df = pd.DataFrame(columns=["Metrics"] + abundance_set + ["Whole dataset"])

    #variable to check if file will be saved
    self.fileSaved = True

    #dictionary to kegg values
    category_to_search = {'KEGG_ko':False, 'KEGG_Pathway':False, 'KEGG_Module':False, 'KEGG_Reaction':False}

    #variable to check connection    
    self.internetWork = True

    #only if online search is request
    if(self.params["keggOnline"]):
      #first of all check for what kegg value is need to search
      for element in self.my_list:
        #by default i put the first element 
        value = element[0]
        #if the element is a aggregation kegg can be only in the second so i put it
        if(len(element) == 2):
          value = element[1]

        #check if the value in the list are one of this to search
        if(value == 'KEGG_ko'):
          category_to_search['KEGG_ko'] = True
        elif(value == 'KEGG_Pathway'):
          category_to_search['KEGG_Pathway'] = True
        elif(value == 'KEGG_Module'):
          category_to_search['KEGG_Module'] = True
        elif(value == 'KEGG_Reaction'):
          category_to_search['KEGG_Reaction'] = True

      #check if is possible get online value of kegg value
      try:
        #try all download here
        if(category_to_search['KEGG_ko']):
          #get online value of KEGG_ko
          self.query_ko = kegg_list("orthology").read()
        if(category_to_search['KEGG_Pathway']):
          #get online value of KEGG_pathway
          self.query_pathway = kegg_list("pathway").read()
        if(category_to_search['KEGG_Module']):
          #get online value of KEGG_module
          self.query_module = kegg_list("module").read()
        if(category_to_search['KEGG_Reaction']):
          #get online value of KEGG_reaction
          self.query_reaction = kegg_list("reaction").read()
      except Exception as e:
        #print("===>>" + str(e))
        self.internetWork = False
        return

      #Manage the online results
      manage_kegg_query(self)

    #for every list element create a file
    for element in self.my_list:
      #get all F cols
      cols = list(self.df.filter(regex=r'F\d+'))
      #add in first place the col_name of column that i want aggragate
      cols.extend(element)
      #add Sequence column to avoid a problem with drop duplicate during the aggregation phase
      #that could remove two row with the same values but of two different sequence
      cols.extend(["Sequence"])
      #create a tmp df
      df_tmp = self.df[cols]

      #First replace empty strings in all columns to mising values:
      df_tmp = df_tmp.replace('', np.nan)

      #prepare filename to save file
      final_path = ""

      #create a tmp df for supplementary tables
      df_tmp_sup = df_tmp.copy()
      #prepare filename to save supplementary tables
      final_path_sup = ""

      #check if there are 1 or 2 name
      if(len(element) == 1):
        #take col name
        col_name = element[0]

        #get abundace colums
        aboundance_cols = list(df_tmp.filter(regex=r'F\d+'))

        #re put nan in empty cells
        df_tmp[aboundance_cols] = df_tmp[aboundance_cols].replace({0:np.nan})

        #controls for supplementary tables
        if(self.params["sup_tab"] or self.params["extra_counts_col"] or self.params["counts_col"]):
          #re put nan in empty cells
          df_tmp_sup[aboundance_cols] = df_tmp_sup[aboundance_cols].replace({0:np.nan})

          #drop unuseless row
          df_tmp_sup = df_tmp_sup.dropna(subset=[col_name])
          #For the safe I convert this column to a string before split
          df_tmp_sup = df_tmp_sup.astype({col_name: 'str'})

          if(self.params["mode"] == "PSMs"):
            df_tmp_sup = (df_tmp_sup.assign(new_col=df_tmp_sup[col_name].str.split('[,;]'))
              .explode('new_col')
              .groupby('new_col', as_index=False)
              .count())
          else: #Proteins/Peptides
            df_tmp_sup = (df_tmp_sup.assign(new_col=df_tmp_sup[col_name].str.split('[,;]'))
              .explode('new_col')
              .drop_duplicates()
              .groupby('new_col', as_index=False)
              .count())

          #edit final_path_sup
          exstension = ""
          col_count = ""
          if(self.params["mode"] == "Proteins"):
            exstension = "-protcounts"
            col_count = "Total protein count"
          elif( (self.params["mode"] == 'Peptides') or (self.params["mode"] == 'PSMs') ):
            exstension = "-peptcounts"
            col_count = "Total peptide count"
          else:
            exstension = "-count"
            col_count = "Total count"

          #raname previus column name to tot
          df_tmp_sup.rename(columns = {col_name:col_count}, inplace = True)
          #rename the tmp col use to explode
          df_tmp_sup.rename(columns = {'new_col':col_name}, inplace = True)
          #remove the new empty values
          df_tmp_sup = df_tmp_sup.replace('', np.nan)
          df_tmp_sup = df_tmp_sup.dropna(subset=[col_name])

          #edit final_path_sup
          final_path_sup = self.params["prefix"] + col_name + exstension + self.params["suffix"]

        #drop unuseless row
        df_tmp = df_tmp.dropna(subset=[col_name])
        #For the safe I convert this column to a string before split
        df_tmp = df_tmp.astype({col_name: 'str'})

        #create the new file with the sum of aboundances
        if(self.params["mode"] == "PSMs"):
          df_tmp = (df_tmp.assign(new_col=df_tmp[col_name].str.split('[,;]'))
            .explode('new_col')
            .groupby('new_col', as_index=False)
            .sum(min_count=1))
        else: #Proteins/Peptides
          #create the new file with the sum of aboundances
          df_tmp = (df_tmp.assign(new_col=df_tmp[col_name].str.split('[,;]'))
            .explode('new_col')
            .drop_duplicates()
            .groupby('new_col', as_index=False)
            .sum(min_count=1))

        #rename the tmp col use to explode
        df_tmp.rename(columns = {'new_col':col_name}, inplace = True)

        #remove the new empty values
        df_tmp = df_tmp.replace('', np.nan)
        df_tmp = df_tmp.dropna(subset=[col_name])

        #edit final_path
        final_path = self.params["prefix"] + col_name + self.params["suffix"]

        ## for df_tmp and df_tmp_sup ##
        #add extra column with description of kegg name
        if( (col_name == "KEGG_ko") and (category_to_search['KEGG_ko']) ):
          #get position for new column
          position_to_insert = df_tmp.columns.get_loc("KEGG_ko")+1
          #crearte a new empty column
          df_tmp.insert(loc=position_to_insert, column="KO name", value=['' for i in range(df_tmp.shape[0])])

          #it looks for all the values of kegg and associates them with the correct description
          for i, row in self.df_ko.iterrows():
            df_tmp["KO name"].where(df_tmp["KEGG_ko"] != row[0], row[1], inplace=True)

          #check for add description also in df_tmp_sup
          if(self.params["sup_tab"]):
            df_tmp_sup.insert(loc=position_to_insert, column="KO name", value=df_tmp["KO name"])

        elif( (col_name == "KEGG_Pathway") and (category_to_search['KEGG_Pathway']) ):
          #get position for new column
          position_to_insert = df_tmp.columns.get_loc("KEGG_Pathway")+1
          #crearte a new empty column
          df_tmp.insert(loc=position_to_insert, column="Pathway name", value=['' for i in range(df_tmp.shape[0])])

          #it looks for all the values of kegg and associates them with the correct description
          for i, row in self.df_pathway.iterrows():
            df_tmp["Pathway name"].where(df_tmp["KEGG_Pathway"] != row[0], row[1], inplace=True)

          #check for add description also in df_tmp_sup
          if(self.params["sup_tab"]):
            df_tmp_sup.insert(loc=position_to_insert, column="Pathway name", value=df_tmp["Pathway name"])

        elif( (col_name == "KEGG_Module") and (category_to_search['KEGG_Module']) ):
          #get position for new column
          position_to_insert = df_tmp.columns.get_loc("KEGG_Module")+1
          #crearte a new empty column
          df_tmp.insert(loc=position_to_insert, column="Module name", value=['' for i in range(df_tmp.shape[0])])

          #it looks for all the values of kegg and associates them with the correct description
          for i, row in self.df_module.iterrows():
            df_tmp["Module name"].where(df_tmp["KEGG_Module"] != row[0], row[1], inplace=True)

          #check for add description also in df_tmp_sup
          if(self.params["sup_tab"]):
            df_tmp_sup.insert(loc=position_to_insert, column="Module name", value=df_tmp["Module name"])

        elif( (col_name == "KEGG_Reaction") and (category_to_search['KEGG_Reaction']) ):
          #get position for new column
          position_to_insert = df_tmp.columns.get_loc("KEGG_Reaction")+1
          #crearte a new empty column
          df_tmp.insert(loc=position_to_insert, column="Reaction name", value=['' for i in range(df_tmp.shape[0])])

          #it looks for all the values of kegg and associates them with the correct description
          for i, row in self.df_reaction.iterrows():
            df_tmp["Reaction name"].where(df_tmp["KEGG_Reaction"] != row[0], row[1], inplace=True)

          #check for add description also in df_tmp_sup
          if(self.params["sup_tab"]):
            df_tmp_sup.insert(loc=position_to_insert, column="Reaction name", value=df_tmp["Reaction name"])

      else:
        #take cols name
        col_name_1 = element[0]
        col_name_2 = element[1]

        #get abundace colums
        aboundance_cols = list(df_tmp.filter(regex=r'F\d+'))

        #re put nan in empty cells
        df_tmp[aboundance_cols] = df_tmp[aboundance_cols].replace({0:np.nan})

        #controls for supplementary tables
        if(self.params["sup_tab"] or self.params["extra_counts_col"] or self.params["counts_col"]):
          #re put nan in empty cells
          df_tmp_sup[aboundance_cols] = df_tmp_sup[aboundance_cols].replace({0:np.nan})

          #drop unuseless row
          df_tmp_sup = df_tmp_sup.dropna(subset=[col_name_1, col_name_2])
          #For the safe I convert this column to a string before split
          df_tmp_sup = df_tmp_sup.astype({col_name_2: 'str'})

          if(self.params["mode"] == "PSMs"):
            df_tmp_sup = (df_tmp_sup.assign(new_col=df_tmp_sup[col_name_2].str.split('[,;]'))
              .explode('new_col')
              .groupby([col_name_1, 'new_col'], as_index=False)
              .count())
          else: #Proteins/Peptides
              df_tmp_sup = (df_tmp_sup.assign(new_col=df_tmp_sup[col_name_2].str.split('[,;]'))
              .explode('new_col')
              .drop_duplicates()
              .groupby([col_name_1, 'new_col'], as_index=False)
              .count())

          #edit final_path_sup
          exstension = ""
          col_count = ""
          if(self.params["mode"] == "Proteins"):
            exstension = "-protcounts"
            col_count = "Total protein count"
          elif( (self.params["mode"] == 'Peptides') or (self.params["mode"] == 'PSMs') ):
            exstension = "-peptcounts"
            col_count = "Total peptide count"
          else:
            exstension = "-count"
            col_count = "Total count"

          #raname previus column name to tot
          df_tmp_sup.rename(columns = {col_name_2:col_count}, inplace = True)
          #rename the tmp col use to explode
          df_tmp_sup.rename(columns = {'new_col':col_name_2}, inplace = True)
          #remove the new empty value
          df_tmp_sup = df_tmp_sup.replace('', np.nan)
          df_tmp_sup = df_tmp_sup.dropna(subset=[col_name_1, col_name_2])

          #edit final_path_sup
          final_path_sup = self.params["prefix"] + col_name_1 + "+" + col_name_2 + exstension + self.params["suffix"]
    
        #drop unuseless row
        df_tmp = df_tmp.dropna(subset=[col_name_1, col_name_2])
        #For the safe I convert this column to a string before split
        df_tmp = df_tmp.astype({col_name_2: 'str'})
        
        #create the new file with the sum of aboundaces
        if(self.params["mode"] == "PSMs"):
          df_tmp = (df_tmp.assign(new_col=df_tmp[col_name_2].str.split('[,;]'))
            .explode('new_col')
            .groupby([col_name_1, 'new_col'], as_index=False)
            .sum(min_count=1))
        else: #Proteins/Peptides
          df_tmp = (df_tmp.assign(new_col=df_tmp[col_name_2].str.split('[,;]'))
            .explode('new_col')
            .drop_duplicates()
            .groupby([col_name_1, 'new_col'], as_index=False)
            .sum(min_count=1))

        #rename the tmp col use to explode
        df_tmp.rename(columns = {'new_col':col_name_2}, inplace = True)

        #remove the new empty value
        df_tmp = df_tmp.replace('', np.nan)
        df_tmp = df_tmp.dropna(subset=[col_name_1, col_name_2])

        #add extra column with description of kegg name
        if( (col_name_2 == "KEGG_ko") and (category_to_search['KEGG_ko']) ):
          #get position for new column
          position_to_insert = df_tmp.columns.get_loc("KEGG_ko")+1
          #crearte a new wmpty column
          df_tmp.insert(loc=position_to_insert, column="KO name", value=['' for i in range(df_tmp.shape[0])])

          #it looks for all the values of kegg and associates them with the correct description
          for i, row in self.df_ko.iterrows():
            df_tmp["KO name"].where(df_tmp["KEGG_ko"] != row[0], row[1], inplace=True)

          #check for add description also in df_tmp_sup
          if(self.params["sup_tab"]):
            df_tmp_sup.insert(loc=position_to_insert, column="KO name", value=df_tmp["KO name"])

        elif( (col_name_2 == "KEGG_Pathway") and (category_to_search['KEGG_Pathway']) ):
          #get position for new column
          position_to_insert = df_tmp.columns.get_loc("KEGG_Pathway")+1
          #crearte a new wmpty column
          df_tmp.insert(loc=position_to_insert, column="Pathway name", value=['' for i in range(df_tmp.shape[0])])

          #it looks for all the values of kegg and associates them with the correct description
          for i, row in self.df_pathway.iterrows():
            df_tmp["Pathway name"].where(df_tmp["KEGG_Pathway"] != row[0], row[1], inplace=True)

          #check for add description also in df_tmp_sup
          if(self.params["sup_tab"]):
            df_tmp_sup.insert(loc=position_to_insert, column="Pathway name", value=df_tmp["Pathway name"])

        elif( (col_name_2 == "KEGG_Module") and (category_to_search['KEGG_Module']) ):
          #get position for new column
          position_to_insert = df_tmp.columns.get_loc("KEGG_Module")+1
          #crearte a new wmpty column
          df_tmp.insert(loc=position_to_insert, column="Module name", value=['' for i in range(df_tmp.shape[0])])

          #it looks for all the values of kegg and associates them with the correct description
          for i, row in self.df_module.iterrows():
            df_tmp["Module name"].where(df_tmp["KEGG_Module"] != row[0], row[1], inplace=True)

          #check for add description also in df_tmp_sup
          if(self.params["sup_tab"]):
            df_tmp_sup.insert(loc=position_to_insert, column="Module name", value=df_tmp["Module name"])

        elif( (col_name_2 == "KEGG_Reaction") and (category_to_search['KEGG_Reaction']) ):
          #get position for new column
          position_to_insert = df_tmp.columns.get_loc("KEGG_Reaction")+1
          #crearte a new wmpty column
          df_tmp.insert(loc=position_to_insert, column="Reaction name", value=['' for i in range(df_tmp.shape[0])])

          #it looks for all the values of kegg and associates them with the correct description
          for i, row in self.df_reaction.iterrows():
            df_tmp["Reaction name"].where(df_tmp["KEGG_Reaction"] != row[0], row[1], inplace=True)

          #check for add description also in df_tmp_sup
          if(self.params["sup_tab"]):
            df_tmp_sup.insert(loc=position_to_insert, column="Reaction name", value=df_tmp["Reaction name"])

        #edit final_path
        final_path = self.params["prefix"] + col_name_1 + "+" + col_name_2 + self.params["suffix"]


      # delete unusless column add only for avoid problem with some duplicate sequence during marge
      if 'Sequence' in df_tmp.columns:
        df_tmp = df_tmp.drop('Sequence', axis=1)
      if 'Sequence' in df_tmp_sup.columns:
        df_tmp_sup = df_tmp_sup.drop('Sequence', axis=1)

      #control to add extra columns in file to show proteins/peptides count
      if self.params["extra_counts_col"] or self.params["counts_col"] :
        # Get last column from df_tmp_sup (that contain counts)
        last_column_df1 = df_tmp_sup.iloc[:, -1]
        # Get the original column name from df_tmp_sup
        column_name = df_tmp_sup.columns[-1]
        # Add last column of df_tmp_sup as last column of df_tmp with original name
        df_tmp[column_name] = last_column_df1

      #control for numbers of proteins/peptides if need
      if self.params["counts_col"] :
        #check to eliminate lines that do not contain enough proteins or peptides
        min_counts = self.params["min_counts"]
        if(min_counts > 0):
          # Delete rows with less than required value in last column
          df_tmp = df_tmp.loc[df_tmp.iloc[:, -1] >= min_counts]
          if self.params["sup_tab"] :
            df_tmp_sup = df_tmp_sup.loc[df_tmp_sup.iloc[:, -1] >= min_counts]

        #check if i need to remove last column from normal file
        #(maybe it was just to remove the subthreshold columns but you didn't want to display them)
        if( not self.params["extra_counts_col"] ):
          # Deleting the last column
          df_tmp = df_tmp.drop(df_tmp.columns[-1], axis=1)

      #control to put zero in empty cells 
      if(self.params['fill0'] == 1):
        #get abundace colums
        sub_set = list(df_tmp.filter(regex=r'F\d+'))
        df_tmp[sub_set] = df_tmp[sub_set].fillna(0)
        if(self.params["sup_tab"]):
          df_tmp_sup[sub_set] = df_tmp_sup[sub_set].fillna(0)

      ### ONLY for Summary metrics ### final_path
      # copy of df
      metrics_df = df_tmp.copy()
      thisName = ""
      #get name and delete useless rows
      if(len(element) == 1):
        #take cols name
        thisName = element[0]
        #delete useless columns
        metrics_df = metrics_df[metrics_df[element[0]] != 'unassigned']
      else:
        #take cols name
        thisName = element[0] +" + "+ element[1]
        #delete useless columns
        metrics_df = metrics_df[metrics_df[element[0]] != 'unassigned']
        metrics_df = metrics_df[metrics_df[element[1]] != 'unassigned']

      ##get correct input type for count e total (summary files)
      column_count_text = ""
      column_total_text = ""
      if(MyUtility.workDict["mode"] == "Proteins"):
        column_count_text = "Quantified"
        column_total_text = "Total abundance"
      elif(MyUtility.workDict["mode"] == "Peptides"):
          column_count_text = "Quantified"
          column_total_text = "Total abundance"
      else:
          column_count_text = "Identified"
          column_total_text = "Total PSMs"

      #conto il totale delle righe che contengono il valore del quale conto le metriche
      whole_count = len(metrics_df)

      # Funzione per calcolare il numero di valori > 0 e diversi da NaN
      def count_positive_notnan(column):
          return (column.gt(0) & ~column.isna()).sum()
      # Calcola il numero di valori per ciascuna colonna
      count_vals = metrics_df[abundance_set].apply(count_positive_notnan)
      # Creo un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
      new_row = {'Metrics': column_count_text+': '+thisName}
      new_row.update(count_vals.to_dict())
      new_row.update({'Whole dataset': whole_count})
      # Aggiungo la nuova riga al DataFrame 'new_df'
      tmp_df = pd.DataFrame(new_row, index=[0])
      #aggiungo al vettore dei risultati
      dfs_to_concat.append(tmp_df)

      # SUM #
      # Funzione per calcolare la somma dei valori > 0 e diversi da NaN
      def sum_positive_notnan(column):
          return column[(column > 0) & (~column.isna())].sum()
      # Calcola la somma per ciascuna colonna
      count_vals = metrics_df[abundance_set].apply(sum_positive_notnan)
      # Creo un dizionario con la nuova riga contenente i nomi delle colonne e i relativi conteggi
      new_row = {'Metrics': column_total_text+': '+thisName}
      new_row.update(count_vals.to_dict())
      #new_row.update({'Whole dataset': whole_count})
      # Aggiungo la nuova riga al DataFrame 'new_df'
      tmp_df = pd.DataFrame(new_row, index=[0])
      #aggiungo al vettore dei risultati
      dfs_to_concat.append(tmp_df)

      #save files
      try:
        #creating final path
        # separa il percorso in base alle barre
        path_list = self.file_path.split("/")
        # separa l'ultima parte del percorso in base al punto
        file_name, file_extension = path_list[-1].rsplit(".", 1)
        # sostituisci la parte compresa tra l'ultimo "/" e l'ultimo "." con la nuova stringa
        final_path = "/".join(path_list[:-1] + [final_path + "." + file_extension])
        # sostituisci la parte compresa tra l'ultimo "/" e l'ultimo "." con la nuova stringa
        final_path_sup = "/".join(path_list[:-1] + [final_path_sup + "." + file_extension])

        #check if path already exist
        alreadyExists = os.path.exists(final_path) or (os.path.exists(final_path_sup) and self.params["sup_tab"])
        if(self.replaceAll == 2):
          if( alreadyExists ):
            msg_box = tk.messagebox.askquestion('Replace or ignore file', 'Some files with the same name already exist in the destination folder.\nReplace them?', icon='warning')
            if msg_box == 'yes':
              self.replaceAll = 1
            else:
              self.replaceAll = 0

        if((self.replaceAll > 0) or (not alreadyExists)):
          #file_path_without_extension, file_extension = self.file.name.rsplit(".", 1)
          if file_extension == "xlsx":
            #save the file
            df_tmp.to_excel(final_path, index=False)
            if(self.params["sup_tab"]):
              df_tmp_sup.to_excel(final_path_sup, index=False)
          else:
            df_tmp.to_csv(final_path, sep='\t', index=False)
            if(self.params["sup_tab"]):
              df_tmp_sup.to_csv(final_path_sup, sep='\t', index=False)
      except Exception as e:
        #print("===>>" + str(e))
        self.fileSaved = False

    #if there is some information than export df
    if(len(dfs_to_concat) > 0):
      # Concatenare tutti i DataFrame nella lista in un unico DataFrame
      metrics_df = pd.concat(dfs_to_concat, ignore_index=True)

      #Riordino le metriche ponendo prima tutti i quantified e poi tutti i total abundance
      # Definisci una funzione per ottenere l'ordine delle metriche
      def get_order(metric):
        if column_count_text in metric:
          return 0
        elif column_total_text in metric:
          return 1
        else:
          return 2  # Se ci sono altre metriche, vengono posizionate alla fine

      # Ordina il DataFrame utilizzando la funzione get_order
      metrics_df['Order'] = metrics_df['Metrics'].apply(get_order)
      metrics_df = metrics_df.sort_values(by=['Order', 'Metrics']).drop('Order', axis=1)

      # Resetta gli indici del DataFrame
      metrics_df = metrics_df.reset_index(drop=True)

      #save df in window
      self.window.metrics_df = metrics_df
    else:
      if( hasattr(self.window, 'metrics_df') ):
        del self.window.metrics_df

#class to rename files
class AsyncRenameFile(Thread):
  def __init__(self, list_file, df_tm):
    super().__init__()

    #save list of file to edit
    self.list_file = list_file
    #save df template
    self.df_tm = df_tm

  def run(self):
    #variable to check if file will be saved
    self.fileSaved = True

    #variable to check correct lenght of all files
    self.correctLen = True
    len_template = self.df_tm[self.df_tm.columns[0]].count()

    #Take the name from template
    col_one_list = self.df_tm['Old Name'].tolist()
    col_two_list = self.df_tm['New Name'].tolist()
    #iterate for every path
    for filepath in self.list_file:
      #open file
      file_extension = filepath.split(".")[-1]
      try:
        if file_extension == "xlsx":
          df = pd.read_excel(filepath)
        else:
          df = pd.read_csv(filepath, sep='\t')
      except Exception as e:
        #print("===>>" + str(e))
        self.fileSaved = False

      #take the old list of "F*" value in order
      cols_len = len(list(df.filter(regex=r'F\d+')))
      
      #check lenght
      if( cols_len == len_template ):

        #Change all name to take only "F+number" for every intressed colums
        df.columns = df.columns.str.replace(r'.*\b(F\d+)\b.*', r'\1', regex=True)

        #the columns are sorted according to the template
        #take the old list of "F*" value in order
        old_cols = list(df.filter(regex=r'F\d+'))
        #calculate name before and after the "f*" coloums
        before_cols = [col for col in df.columns if df.columns.get_loc(col) < df.columns.get_loc(old_cols[0])]
        after_cols = [col for col in df.columns if df.columns.get_loc(col) > df.columns.get_loc(old_cols[-1])]
        #reorder df according to the new order
        df = df[before_cols+col_one_list+after_cols]      

        for i in range(len(col_one_list)):
          df.rename(columns = {col_one_list[i]:col_two_list[i]}, inplace = True)

        #check to save file  
        try:
          #file_path_without_extension, file_extension = self.file.name.rsplit(".", 1)
          if file_extension == "xlsx":
            df.to_excel(filepath, index=False)
          else:
            df.to_csv(filepath, sep='\t', index=False, header=True,)
        except Exception as e:
          #print("===>>" + str(e))
          self.fileSaved = False
      else:
        self.correctLen = False
      
#class to manage file on protein window
#changing the order of the controls inside the function can break the correct functioning##
class ManageData(Thread):
  def __init__(self, window):
    super().__init__()

    #take window data
    self.window = window

  def run(self):
    #take a copy of window to do controls
    window = self.window
    #create a copy for finale edits
    df_final = window.df.copy()

    #function for concatenate values
    def join_unique(x):
      if len(set(x)) == 1:
          return x.iloc[0]
      else:
          return '; '.join(set(x))

    #Solve problem with row that contains the same sequence value
    if( (MyUtility.workDict["input_type"] == 'mzTab') and (MyUtility.workDict["mode"] == 'Peptides') ):
      #take the aboundance sub set
      sub_set = list(df_final.filter(regex=r'Abundance F\d+'))
      #create a list with a column list that need to be equal in first phase
      column_list = ['Sequence']+sub_set

      #PHASE_1# Group by if all columns in column_list are equal and join the different Accession
      df_final = df_final.groupby(column_list).agg({'Master Protein Accessions': join_unique}).reset_index()
      
      #PHASE2# Sum row whit same Sequence
      agg_dict = {name: 'sum' for name in sub_set}
      df_final = df_final.groupby('Sequence').agg({**agg_dict, **{col: 'first' for col in df_final.columns if col not in agg_dict}}).set_index('Sequence')
      df_final = df_final.reset_index()

      #put Accession in second coulumn
      all_cols = list(df_final.columns)
      all_cols.insert(1, all_cols.pop(all_cols.index('Master Protein Accessions')))
      df_final = df_final.reindex(columns=all_cols)

    #control for Protein FDR (Confidence)
    if( hasattr(window, 'frame_confidence') and (window.frame_confidence.grid_info() != {}) ):
      if(MyUtility.workDict['mode'] == 'Proteins'):
        condidence_column_name = 'Protein FDR Confidence: Combined'
      else:
        condidence_column_name = 'Confidence'
      if(window.var_chc_low.get() == 0 ):
        df_final.drop(df_final.index[df_final[condidence_column_name] == 'Low'], inplace=True)
      if(window.var_chc_medium.get() == 0 ):
        df_final.drop(df_final.index[df_final[condidence_column_name] == 'Medium'], inplace=True)
      if(window.var_chc_high.get() == 0 ):
        df_final.drop(df_final.index[df_final[condidence_column_name] == 'High'], inplace=True)
    
    #control for normalized
    if( hasattr(window, 'chc_normalized') and (window.chc_normalized.grid_info() != {}) ):
      if(window.var_chc_normalized.get() == 1):
        df_final.drop(list(df_final.filter(regex = 'Abundance:')), axis = 1, inplace = True)
      else:
        df_final.drop(list(df_final.filter(regex = 'Normalized')), axis = 1, inplace = True)

    #After normalized control rename columns name
    if(MyUtility.workDict["input_type"] == 'proteome'):
      def rename_columns(col):
        if col.startswith('Abundances (Normalized): F'):
            new_col = f"Abundance {col.split(' ')[2]}"
            new_col = new_col[:-1]
        elif col.startswith('Abundance: F'):
            new_col = f"Abundance {col.split(' ')[1]}"
            new_col = new_col[:-1]
        else:
            new_col = col
        return new_col

      df_final = df_final.rename(columns=lambda col: rename_columns(col))

    #Control for description
    if( hasattr(window, 'frame_description') and (window.frame_description.grid_info() != {}) ):
      #find the correct name for the filtre
      if(MyUtility.workDict["mode"] == 'Proteins'):
        description_name = 'Description'
      else:
        description_name = 'Master Protein Descriptions'

      #get content of description listbox
      get_content = window.dsc_listbox.get(0, END)
      #make list for remove row
      toSearch = []
      for con_item in get_content:
        toSearch.append(con_item)
      #delete rows that do not contain any words in the column
      if(window.rdb_var.get()=='or'):
        df_final = df_final[df_final[description_name].str.contains('|'.join(toSearch)) == True]
      elif(window.rdb_var.get()=='and'):
        base = r'^{}'
        expr = '(?=.*{})'
        toSearch = base.format(''.join(expr.format(w) for w in toSearch))
        df_final = df_final[df_final[description_name].str.contains(toSearch) == True]
      #finally edit the cells for remove "newline"(\n) and put ";"
      df_final[description_name] = df_final[description_name].str.replace("\n","; ")
    
    #control for Master Protein
    if( hasattr(window, 'chc_master')):
      if(window.var_chc_master.get() == 1):
        df_final.drop(df_final.index[df_final['Master'] != 'Master Protein'], inplace=True)
    
    #control for Protein Accessions
    if( hasattr(window, 'chc_ptrAccessions')):
      if(window.var_chc_ptrAccessions.get() == 0):
        df_final.drop(['Protein Accessions'], inplace=True, axis=1, errors='ignore')
    
    #control for marked as (Marker)
    if( hasattr(window, 'frame_marker') and (window.frame_marker.grid_info() != {}) ):
      if('Marked as' in df_final.columns):
        i = 0
        for marked in window.scl_check_marker.chcs:
          if(window.scl_check_marker.var_chcs[i].get() == 0):
            if(marked.cget("text") == "Empty"):
              df_final = df_final.dropna(subset=['Marked as'])
            else:
              #print(marked.cget("text"))
              df_final.drop(df_final.index[df_final['Marked as'] == marked.cget("text")], inplace=True)
          i = i+1
    
    #control for Quan Info
    if( hasattr(window, 'frame_quanInfo') and (window.frame_quanInfo.grid_info() != {}) ):
      i = 0
      for quan in window.scl_check_quantInfo.chcs:
        if(window.scl_check_quantInfo.var_chcs[i].get() == 0):
          df_final.drop(df_final.index[df_final['Quan Info'] == quan.cget("text")], inplace=True)
        i = i+1

      #get abundace colums
      sub_set = list(df_final.filter(regex=r'F\d+'))
    
    #Control only for peptide on Protoemoe Discovery
    if( (MyUtility.workDict["input_type"] == 'proteome') and (MyUtility.workDict["mode"] == 'Peptides') ):
      #remove duplicate Sequence (cause of Protoeme Discovery bug)
      #create a list for Abundance cols
      abn_list = list(df_final.filter(regex=r'F\d+'))
      #create a list for all cols
      c_list = list(df_final.columns)
      #remove Sequence (because used for groupby)
      c_list.remove('Sequence')
      #remove all Abundance (because used for sum)
      c_list = [i for i in c_list if i not in abn_list]
      #aggregate by sequence and sum Abundance
      df_final = df_final.groupby('Sequence').agg({**dict.fromkeys(c_list, 'first'), **dict.fromkeys(abn_list, 'sum') }).reset_index()
      #re-swap "Sequence"(first columns) and "Confidence"(second columns)
      # get a list of the columns
      col_list = list(df_final)
      # use this handy way to swap the elements
      col_list[0], col_list[1] = col_list[1], col_list[0]
      # assign back, the order will now be swapped
      df_final = df_final[col_list]
      #re put nan in empty cells
      df_final[abn_list] = df_final[abn_list].replace({0:np.nan})

      #show at first Sequence and at second Master Protein Accessions
      #extract columns from dataframe
      sequence_col = df_final.pop('Sequence')
      master_col = df_final.pop('Master Protein Accessions')
      #insert columns in correct order
      df_final.insert(0, 'Sequence', sequence_col)
      df_final.insert(1, 'Master Protein Accessions', master_col)
    
    #create abundace colums after drop some
    sub_set = list(df_final.filter(regex=r'F\d+'))

    #if we are on mzTab, convert string in integer for abundance
    if(MyUtility.workDict["input_type"] == 'mzTab' ):
      df_final[sub_set] = df_final[sub_set].apply(pd.to_numeric, errors='coerce')

    #control for abundance
    if( hasattr(window, 'frame_validValues') and (window.frame_validValues.grid_info() != {}) ):
      num_abundance = 0
      if(window.opt_abundance_var.get() == 'Absolute'):
        num_abundance = int(window.ntr_abundance.get())
      elif(window.opt_abundance_var.get() == 'Percentage'):
        #get number of columns
        num_cols = len(sub_set)
        #calcolate num
        num_abundance = window.proper_round((int(window.ntr_abundance.get()) * num_cols)/100)
      #delete all row width 'num_cols' empty in Aboundance(F1,F2..) columns
      df_final = df_final.dropna(subset=sub_set, thresh=num_abundance)
    
    #recreate abundace colums after drop some
    sub_set = list(df_final.filter(regex=r'F\d+'))

    #control for Re-Normalized
    if( hasattr(window, 'chc_re_normalized')  and (window.chc_re_normalized.grid_info() != {}) ):
      if(window.var_chc_re_normalized.get() == 1):
        #RE-Normalize all colums
        for col_name in sub_set:
          df_final[col_name] = ( df_final[col_name]/df_final[col_name].sum() ) * 10000000000

    #Control only for PSMs on Protoemoe Discovery
    if( (MyUtility.workDict["input_type"] == 'proteome') and (MyUtility.workDict["mode"] == 'PSMs') ):
      #prepare pivot table
      table = pd.pivot_table(data=df_final,
                             index=['Sequence'],
                             columns=['File ID'],
                             aggfunc='size')
                             #fill_value=0) if I want fill Nan with 0
      
      #table.columns = table.columns.droplevel(0) #remove first line
      table.columns.name = None               #File id
      table = table.reset_index()                #index to columns

      #marge original file with table
      df_final = df_final.merge(table, left_on='Sequence', right_on='Sequence', how='left')

      #remove duplicate
      df_final = df_final.drop_duplicates(subset='Sequence', keep="last")

      #remove "File ID" column
      df_final.drop(['File ID'], inplace=True, axis=1, errors='ignore')

      #reorder F columns
      num = df_final.columns.str.extract('F(\d+)', expand=False).astype(float)
      cols = df_final.columns.to_numpy(copy=True)
      m = num.notna()
      order = np.argsort(num[m])
      cols[m] = cols[m][order]
      df_final = df_final[cols]
      
    #Control only for PSMs on mzTab
    if( (MyUtility.workDict["input_type"] == 'mzTab') and (MyUtility.workDict["mode"] == 'PSMs') ):
      #PHASE_1# Group by if all columns in column_list are equal and join the different Accession
      df_final = df_final.groupby('PSM_ID').agg({'Master Protein Accessions': join_unique, **{col: 'first' for col in df_final.columns if col not in ['PSM_ID','Master Protein Accessions']} }).reset_index()

      #remove unuse part of file
      df_final['Spectra_ref'] = 'F' + df_final['Spectra_ref'].str.extract(r'\[(\d+)\]')
    
      #prepare pivot table
      table = pd.pivot_table(data=df_final,
                             index=['Sequence'],
                             columns=['Spectra_ref'],
                             aggfunc='size')
                             #fill_value=0) if I want fill Nan with 0
      
      #table.columns = table.columns.droplevel(0) #remove first line
      table.columns.name = None               #File id
      table = table.reset_index()                #index to columns

      #marge original file with table
      df_final = df_final.merge(table, left_on='Sequence', right_on='Sequence', how='left')

      #remove duplicate
      df_final = df_final.drop_duplicates(subset='Sequence', keep="last")

      #remove "Spectra_ref"
      df_final.drop(['Spectra_ref'], inplace=True, axis=1, errors='ignore')
      #remove "File ID" column
      df_final.drop(['PSM_ID'], inplace=True, axis=1, errors='ignore')

      #reorder F columns
      num = df_final.columns.str.extract('F(\d+)', expand=False).astype(float)
      cols = df_final.columns.to_numpy(copy=True)
      m = num.notna()
      order = np.argsort(num[m])
      cols[m] = cols[m][order]
      df_final = df_final[cols]

      #show at first Sequence and at second Master Protein Accessions
      #extract columns from dataframe
      sequence_col = df_final.pop('Sequence')
      master_col = df_final.pop('Master Protein Accessions')
      #insert columns in correct order
      df_final.insert(0, 'Sequence', sequence_col)
      df_final.insert(1, 'Master Protein Accessions', master_col)
      
    #control to put zero in empty cells
    if( hasattr(window, 'chc_fill_zero') and (window.chc_fill_zero.grid_info() != {}) ):
      if(window.var_chc_fill_zero.get() == 1):
        df_final[sub_set] = df_final[sub_set].fillna(0)
    
    #Final Reorder
    if( (MyUtility.workDict["input_type"] == 'proteome') or (MyUtility.workDict["input_type"] == 'mzTab') ):
      if(MyUtility.workDict["mode"] == 'Proteins'):
        #before save, reorder file according to "Accession" column
        df_final = df_final.sort_values('Accession')
      elif(MyUtility.workDict["mode"] == 'Peptides'):
        #before save, reorder file according to "Sequence" column
        df_final = df_final.sort_values('Sequence')
      elif(MyUtility.workDict["mode"] == 'PSMs'):
        #before save, reorder file according to "Sequence" column
        df_final = df_final.sort_values('Sequence')


    #save edit df in tmp variable
    window.df_tmp = df_final

class ManageDataDynamic(Thread):
  def __init__(self, window):
    super().__init__()

    #take window data
    self.window = window

  def run(self):
    #take a copy of window to do controls
    window = self.window
    #create a copy for finale edits
    df_final = window.df.copy()

    #Control for rename columns
    #Protein Acession
    if( hasattr(window, 'frame_proteinAccession') ):
      old_name = window.proteinAccession_listbox.get(0, tk.END)[0]
      # Rinominiamo la colonna
      if(MyUtility.workDict['mode'] == 'Proteins'):
        df_final = df_final.rename(columns={old_name: 'Accession'})
      else:
        df_final = df_final.rename(columns={old_name: 'Master Protein Accessions'})
    #Peptide Sequence
    if( hasattr(window, 'frame_peptideSequence') ):
      old_name = window.peptideSequence_listbox.get(0, tk.END)[0]
      # Rinominiamo la colonna
      df_final = df_final.rename(columns={old_name: 'Sequence'})
    #Sample ID
    if(hasattr(window, 'frame_sampleID')):
      old_name = window.sampleID_listbox.get(0, tk.END)[0]
      df_final = df_final.rename(columns={old_name: 'File ID'})
    #Abundance
    if(hasattr(window, 'frame_abundanceValue')):
      for index in range(window.abundanceValue_listbox.size()):
        old_name = window.abundanceValue_listbox.get(index)
        new_name = 'Abundance F'+str(index+1) #put index+1 because we want count starting from 1 and not 0
        df_final = df_final.rename(columns={old_name: new_name})

    #Remove useless columns
    #get list of unused colums
    useless_column = window.all_columns_listbox.get(0, tk.END)
    existing_columns = [col for col in useless_column if col in df_final.columns]
    df_final = df_final.drop(existing_columns, axis=1)

    #create abundace colums after drop some
    sub_set = list(df_final.filter(regex=r'F\d+'))

    #Convert string in integer for abundance
    df_final[sub_set] = df_final[sub_set].apply(pd.to_numeric, errors='coerce')

    #control for abundance
    if( hasattr(window, 'frame_validValues') and (window.frame_validValues.grid_info() != {}) ):
      num_abundance = 0
      if(window.opt_abundance_var.get() == 'Absolute'):
        num_abundance = int(window.ntr_abundance.get())
      elif(window.opt_abundance_var.get() == 'Percentage'):
        #get number of columns
        num_cols = len(sub_set)
        #calcolate num
        num_abundance = window.proper_round((int(window.ntr_abundance.get()) * num_cols)/100)
      #delete all row width 'num_cols' empty in Aboundance(F1,F2..) columns
      df_final = df_final.dropna(subset=sub_set, thresh=num_abundance)
    
    #recreate abundace colums after drop some
    sub_set = list(df_final.filter(regex=r'F\d+'))

    #control for Re-Normalized
    if( hasattr(window, 'chc_re_normalized')  and (window.chc_re_normalized.grid_info() != {}) ):
      if(window.var_chc_re_normalized.get() == 1):
        #RE-Normalize all colums
        for col_name in sub_set:
          df_final[col_name] = ( df_final[col_name]/df_final[col_name].sum() ) * 10000000000


    #Control only for PSMs on Protoemoe Discovery
    if(MyUtility.workDict["mode"] == 'PSMs'):
      #prepare pivot table
      table = pd.pivot_table(data=df_final,
                             index=['Sequence'],
                             columns=['File ID'],
                             aggfunc='size')
                             #fill_value=0) if I want fill Nan with 0
      
      #table.columns = table.columns.droplevel(0) #remove first line
      table.columns.name = None               #File id
      table = table.reset_index()                #index to columns

      #marge original file with table
      df_final = df_final.merge(table, left_on='Sequence', right_on='Sequence', how='left')

      #remove duplicate
      df_final = df_final.drop_duplicates(subset='Sequence', keep="last")

      #remove "File ID" column
      df_final.drop(['File ID'], inplace=True, axis=1, errors='ignore')

      #reorder F columns
      num = df_final.columns.str.extract('F(\d+)', expand=False).astype(float)
      cols = df_final.columns.to_numpy(copy=True)
      m = num.notna()
      order = np.argsort(num[m])
      cols[m] = cols[m][order]
      df_final = df_final[cols]
     
    #control to put zero in empty cells
    if( hasattr(window, 'chc_fill_zero') and (window.chc_fill_zero.grid_info() != {}) ):
      if(window.var_chc_fill_zero.get() == 1):
        df_final[sub_set] = df_final[sub_set].fillna(0)

    #Final Reorder
    if(MyUtility.workDict["mode"] == 'Proteins'):
      #before save, reorder file according to "Accession" column
      df_final = df_final.sort_values('Accession')
    elif(MyUtility.workDict["mode"] == 'Peptides'):
      #before save, reorder file according to "Sequence" column
      df_final = df_final.sort_values('Sequence')
    elif(MyUtility.workDict["mode"] == 'PSMs'):
      #before save, reorder file according to "Sequence" column
      df_final = df_final.sort_values('Sequence')

    #save edit df in tmp variable
    window.df_tmp = df_final
    
#class to manage file on Taxonomic window
class ManageTaxonomic(Thread):
  def __init__(self, window):
    super().__init__()

    #take window data
    self.window = window

  def run(self):
    #take a copy of window to do controls
    window = self.window
    #create a copy for finale edits
    df_final = window.df.copy()
    df_final_annotation = window.df_annotation.copy()

    #if it is necessary to rename the columns to avoid that in the main
    #df and inn the one with the ci notations there are duplicates
    rename_dict = {col: col+'_2' for col in df_final_annotation.columns if col in df_final.columns}
    df_final_annotation = df_final_annotation.rename(columns=rename_dict)

    #check if the I=L checkbox exist
    isI_equal_L = 0
    if( hasattr(window, 'var_chc_IandL') ):
      isI_equal_L = window.var_chc_IandL.get()

    #if mode is not proteins and I=L is selected we need to add one columns
    if( (MyUtility.workDict["mode"] != 'Proteins') and (isI_equal_L == 1)):
      #add duplicate of sequence column
      #make copy of Sequence column
      column_to_duplicate = df_final['Sequence']
      #get the position of the original sequence column 
      position_to_insert = df_final.columns.get_loc("Sequence")+1
      #put new duplicate column next to original sequence
      df_final.insert(position_to_insert, "Sequence(I=L)", column_to_duplicate)
      #change all I with L
      df_final['Sequence(I=L)'] = df_final['Sequence(I=L)'].replace('I','L', regex=True)

    #merge files
    if(MyUtility.workDict["mode"] == 'Proteins'):
      colname = df_final_annotation.columns[0]
      df_final = df_final.merge(df_final_annotation, left_on='Accession', right_on=colname, how='left')
      df_final.drop(['Accession No.'], inplace=True, axis=1, errors='ignore')
    else:
      if(isI_equal_L == 1):
        df_final = df_final.merge(df_final_annotation, left_on='Sequence(I=L)', right_on='peptide', how='left')
        df_final.drop(['peptide'], inplace=True, axis=1, errors='ignore')
      else:
        df_final = df_final.merge(df_final_annotation, left_on='Sequence', right_on='peptide', how='left')
        df_final.drop(['peptide'], inplace=True, axis=1, errors='ignore')


    #control to put zero in empty cells 
    if(MyUtility.workDict['fill0'] == 1):
      #get abundace colums
      sub_set = list(df_final.filter(regex=r'F\d+'))
      df_final[sub_set] = df_final[sub_set].fillna(0)

    #check if fill empty cells in annotation
    if(window.var_chc_unassigned.get() == 1):
      # Verifica la presenza delle colonne di interesse nel DataFrame
      all_columns_to_fill = df_final_annotation.columns.tolist()
      columns_to_fill = [col for col in all_columns_to_fill if col in df_final.columns.tolist()]
      df_final[columns_to_fill] = df_final[columns_to_fill].fillna(value='unassigned')
      # Sostituisci le stringhe vuote ('') con "unassigned" nelle colonne di interesse
      df_final[columns_to_fill] = df_final[columns_to_fill].replace('', 'unassigned')


    #Add table to dict for aggregation windows
    MyUtility.workDict['taxonomic_table'] = ["superkingdom", "phylum", "class", "order", "family", "genus", "species"]


    #save edit df in tmp variable
    window.df_tmp = df_final

class ManageTaxonomicDynamic(Thread):
  def __init__(self, window):
    super().__init__()

    #take window data
    self.window = window

  def run(self):
    #take a copy of window to do controls
    window = self.window
    #create a copy for finale edits
    df_final = window.df.copy()
    df_final_annotation = window.df_annotation.copy()

    #Protein Acession
    if( hasattr(window, 'frame_proteinAccession') ):
      old_name = window.proteinAccession_listbox.get(0, tk.END)[0]
      # Rinominiamo la colonna
      df_final_annotation = df_final_annotation.rename(columns={old_name: 'proteinAccession'})
    #Peptide Sequence
    if( hasattr(window, 'frame_peptideSequence') ):
      old_name = window.peptideSequence_listbox.get(0, tk.END)[0]
      # Rinominiamo la colonna
      df_final_annotation = df_final_annotation.rename(columns={old_name: 'peptideSequence'})

    #Remove useless columns
    #get list of unused colums
    useless_column = window.all_columns_listbox.get(0, tk.END)
    existing_columns = [col for col in useless_column if col in df_final_annotation.columns]
    df_final_annotation = df_final_annotation.drop(existing_columns, axis=1)

    #if it is necessary to rename the columns to avoid that in the main
    #df and inn the one with the ci notations there are duplicates
    rename_dict = {col: col+'_2' for col in df_final_annotation.columns if col in df_final.columns}
    df_final_annotation = df_final_annotation.rename(columns=rename_dict)

    #check if the I=L checkbox exist
    isI_equal_L = 0
    if( hasattr(window, 'var_chc_IandL') ):
      isI_equal_L = window.var_chc_IandL.get()

    #if mode is not proteins and I=L is selected we need to add one columns
    if( (MyUtility.workDict["mode"] != 'Proteins') and (isI_equal_L == 1)):
      #add duplicate of sequence column
      #make copy of Sequence column
      column_to_duplicate = df_final['Sequence']
      #get the position of the original sequence column 
      position_to_insert = df_final.columns.get_loc("Sequence")+1
      #put new duplicate column next to original sequence
      df_final.insert(position_to_insert, "Sequence(I=L)", column_to_duplicate)
      #change all I with L
      df_final['Sequence(I=L)'] = df_final['Sequence(I=L)'].replace('I','L', regex=True)

    #merge files
    if(MyUtility.workDict["mode"] == 'Proteins'):
      df_final = df_final.merge(df_final_annotation, left_on='Accession', right_on='proteinAccession', how='left')
      df_final.drop(['proteinAccession'], inplace=True, axis=1, errors='ignore')
    else:
      if(MyUtility.workDict['taxonomic_match'] == 'protein'):
        df_final = df_final.merge(df_final_annotation, left_on='Master Protein Accessions', right_on='proteinAccession', how='left')
        df_final.drop(['proteinAccession'], inplace=True, axis=1, errors='ignore')
      else: #peptide
        if(isI_equal_L == 1):
          df_final = df_final.merge(df_final_annotation, left_on='Sequence(I=L)', right_on='peptideSequence', how='left')
          df_final.drop(['peptideSequence'], inplace=True, axis=1, errors='ignore')
        else:
          df_final = df_final.merge(df_final_annotation, left_on='Sequence', right_on='peptideSequence', how='left')
          df_final.drop(['peptideSequence'], inplace=True, axis=1, errors='ignore')

    #control to put zero in empty cells 
    if(MyUtility.workDict['fill0'] == 1):
      #get abundace colums
      sub_set = list(df_final.filter(regex=r'F\d+'))
      df_final[sub_set] = df_final[sub_set].fillna(0)

    #check if fill empty cells in annotation
    if(window.var_chc_unassigned.get() == 1):
      # Verifica la presenza delle colonne di interesse nel DataFrame
      all_columns_to_fill = df_final_annotation.columns.tolist()
      columns_to_fill = [col for col in all_columns_to_fill if col in df_final.columns.tolist()]
      df_final[columns_to_fill] = df_final[columns_to_fill].fillna(value='unassigned')
      # Sostituisci le stringhe vuote ('') con "unassigned" nelle colonne di interesse
      df_final[columns_to_fill] = df_final[columns_to_fill].replace('', 'unassigned')

    #Add table to dict for aggregation windows
    MyUtility.workDict['taxonomic_table'] = [col for col in df_final_annotation.columns if col not in ['Accession', 'Sequence', 'proteinAccession', 'peptideSequence']]

    #save edit df in tmp variable
    window.df_tmp = df_final

#class to manage file on Functional window
class ManageFunctional(Thread):
  def __init__(self, window):
    super().__init__()

    #take window data
    self.window = window

  def run(self):
    #take a copy of window to do controls
    window = self.window
    #create a copy for finale edits
    df_final = window.df.copy()
    df_final_annotation = window.df_annotation.copy()

    #eventual edit for dynamic
    if(MyUtility.workDict["functional_mode"] == 'dynamic'):
      #Protein Acession
      if( hasattr(window, 'frame_proteinAccession') ):
        old_name = window.proteinAccession_listbox.get(0, tk.END)[0]
        # Rinominiamo la colonna
        df_final_annotation = df_final_annotation.rename(columns={old_name: 'query'})
      if( hasattr(window, 'frame_peptideSequence') ):
        old_name = window.peptideSequence_listbox.get(0, tk.END)[0]
        # Rinominiamo la colonna
        df_final_annotation = df_final_annotation.rename(columns={old_name: 'query'})
      #KEGG
      if( hasattr(window, 'frame_kegg_ko') and (window.kegg_ko_listbox.size() > 0) ):
        old_name = window.kegg_ko_listbox.get(0, tk.END)[0]
        # Rinominiamo la colonna
        df_final_annotation = df_final_annotation.rename(columns={old_name: 'KEGG_ko'})
      if( hasattr(window, 'frame_kegg_pathway') and (window.kegg_pathway_listbox.size() > 0) ):
        old_name = window.kegg_pathway_listbox.get(0, tk.END)[0]
        # Rinominiamo la colonna
        df_final_annotation = df_final_annotation.rename(columns={old_name: 'KEGG_Pathway'})
      if( hasattr(window, 'frame_kegg_module') and (window.kegg_module_listbox.size() > 0) ):
        old_name = window.kegg_module_listbox.get(0, tk.END)[0]
        # Rinominiamo la colonna
        df_final_annotation = df_final_annotation.rename(columns={old_name: 'KEGG_Module'})
      if( hasattr(window, 'frame_kegg_reaction') and (window.kegg_reaction_listbox.size() > 0) ):
        old_name = window.kegg_reaction_listbox.get(0, tk.END)[0]
        # Rinominiamo la colonna
        df_final_annotation = df_final_annotation.rename(columns={old_name: 'KEGG_Reaction'})
      if( hasattr(window, 'frame_cog') and (window.cog_listbox.size() > 0) ):
        old_name = window.cog_listbox.get(0, tk.END)[0]
        # Rinominiamo la colonna
        df_final_annotation = df_final_annotation.rename(columns={old_name: 'COG_category'})

      #Remove useless columns
      #get list of unused colums
      useless_column = window.all_columns_listbox.get(0, tk.END)
      existing_columns = [col for col in useless_column if col in df_final_annotation.columns]
      df_final_annotation = df_final_annotation.drop(existing_columns, axis=1)

    #If the cog column is present inside the dataframe, I separate all the letters inside it with a comma
    if 'COG_category' in df_final_annotation.columns:
      # Applying the function to strings with more than one character
      df_final_annotation['COG_category'] = np.where(df_final_annotation['COG_category'].str.len() > 1,
                                 df_final_annotation['COG_category'].str.replace('', ',').str[1:-1], #avendo le virgole le rimuovo
                                 df_final_annotation['COG_category'])
      # Finding single-character strings
      single_char_mask = df_final_annotation['COG_category'].str.len() == 1
      # Replacing single-character strings with the original value
      df_final_annotation.loc[single_char_mask, 'COG_category'] = df_final_annotation.loc[single_char_mask, 'COG_category'].str[0]


    #if it is necessary to rename the columns to avoid that in the main
    #df and inn the one with the ci notations there are duplicates
    rename_dict = {col: col+'_2' for col in df_final_annotation.columns if col in df_final.columns}
    df_final_annotation = df_final_annotation.rename(columns=rename_dict)

    #remove all "-" in the cells
    df_final.replace("-","",inplace=True)
    df_final_annotation.replace("-","",inplace=True)

    #check if fill empty cells in annotation
    if(window.var_chc_unassigned.get() == 1):
      #extra check for only string
      df_final_annotation = df_final_annotation.applymap(lambda x: 'unassigned' if pd.isna(x) or (isinstance(x, str) and x.strip() == '') else x)

    #check if the I=L checkbox exist
    isI_equal_L = 0
    if( hasattr(window, 'var_chc_IandL') ):
      isI_equal_L = window.var_chc_IandL.get()

    #if mode is not proteins and I=L is selected we need to add one columns
    if( (MyUtility.workDict["mode"] != 'Proteins') and (isI_equal_L == 1) and ('Sequence(I=L)' not in df_final.columns)):
      #add duplicate of sequence column
      #make copy of Sequence column
      column_to_duplicate = df_final['Sequence']
      #get the position of the original sequence column 
      position_to_insert = df_final.columns.get_loc("Sequence")+1
      #put new duplicate column next to original sequence
      df_final.insert(position_to_insert, "Sequence(I=L)", column_to_duplicate)
      #change all I with L
      df_final['Sequence(I=L)'] = df_final['Sequence(I=L)'].replace('I','L', regex=True)

    #Check if i need to take information starting from protein file or another
    if(MyUtility.workDict['mode'] == 'Proteins'):
      columns_to_match = 'Accession'
    else:
      if(MyUtility.workDict['functional_match'] == 'protein'):
        columns_to_match = 'Master Protein Accessions'
      else: #peptide
        columns_to_match = 'Sequence'

    #join final_df and annotation_df
    df_final = (df_final.assign(query = df_final[columns_to_match].str.split('; '))
             .explode('query')
             .reset_index()
             .merge(df_final_annotation, how='left', on='query')
             .fillna('')
             .astype(dict.fromkeys(df_final_annotation, str))
             .groupby('index')
             .agg({**dict.fromkeys(df_final, 'first'), **dict.fromkeys(df_final_annotation, ';'.join)})
             .rename_axis(None))
    
    ## Il seguente codice serve per sostituire le posizioni vuote con "unassigned"
    # Converti tutte le colonne in stringhe prima di applicare la funzione
    df_final[df_final_annotation.columns] = df_final[df_final_annotation.columns].astype(str)
    
    #remove a unused coloum (to do after the previuos check on df_final_annotation.columns)
    df_final.drop('query', axis=1, inplace=True)

    # Definisci una funzione per aggiungere 'Z' secondo le tue regole
    def add_unassigned_to_void(cell):
        if cell.startswith(';'):
            cell = 'unassigned' + cell
        if cell.endswith(';'):
            cell = cell + 'unassigned'
        while ";;" in cell:
            cell = cell.replace(";;", ";unassigned;")
        return cell

    #check if fill empty cells in annotation
    if(window.var_chc_unassigned.get() == 1):
      # Applica la funzione alle colonne desiderate in df_final
      for col in df_final_annotation.columns:
          if col in df_final.columns:
              df_final[col] = df_final[col].apply(add_unassigned_to_void)
    ##
    #Replace column values if it repeats the same ";" character
    df_final = df_final.mask(df_final.applymap(lambda x: isinstance(x, str) and set(x) == {';'}), '')
    #other method
    #cols = ['col_1','col_2']
    #df[cols] = df[cols].replace(r'^;{1,}$','', regex=True)

    #control to put zero in empty cells 
    if(MyUtility.workDict['fill0'] == 1):
      #get abundace colums
      sub_set = list(df_final.filter(regex=r'F\d+'))
      df_final[sub_set] = df_final[sub_set].fillna(0)

    #before control for kegg, remove 'ko:' from KEGG_KO
    if 'KEGG_ko' in df_final.columns:
      df_final['KEGG_ko'] = df_final['KEGG_ko'].str.replace('ko:', '')

    #replace empty cell with "unassigned"
    #check if fill empty cells in annotation
    if(window.var_chc_unassigned.get() == 1):
      # Verifica la presenza delle colonne di interesse nel DataFrame
      all_columns_to_fill = df_final_annotation.columns.tolist()
      columns_to_fill = [col for col in all_columns_to_fill if col in df_final.columns.tolist()]
      df_final[columns_to_fill] = df_final[columns_to_fill].fillna(value='unassigned')
      # Sostituisci le stringhe vuote ('') con "unassigned" nelle colonne di interesse
      df_final[columns_to_fill] = df_final[columns_to_fill].replace('', 'unassigned')

    #Control to get KEGG data
    if(window.var_chc_kegg_description.get() == 1):
      #Variable to monitorize internet work
      self.internetWork = True
      #check if is possible get online value of kegg value
      try:
        #try all download here
        if(MyUtility.workDict['functional_mode'] == 'dynamic'):
          #get online value of KEGG_ko
          if window.kegg_ko_listbox.size() > 0:
            self.query_ko = kegg_list("orthology").read()
          #get online value of KEGG_pathway
          if window.kegg_pathway_listbox.size() > 0:
            self.query_pathway = kegg_list("pathway").read()
          #get online value of KEGG_module
          if window.kegg_module_listbox.size() > 0:
            self.query_module = kegg_list("module").read()
          #get online value of KEGG_reaction
          if window.kegg_reaction_listbox.size() > 0:
            self.query_reaction = kegg_list("reaction").read()
        else:
          #get online value of KEGG_ko
          self.query_ko = kegg_list("orthology").read()
          #get online value of KEGG_pathway
          self.query_pathway = kegg_list("pathway").read()
          #get online value of KEGG_module
          self.query_module = kegg_list("module").read()
          #get online value of KEGG_reaction
          self.query_reaction = kegg_list("reaction").read()

      except Exception as e:
        #print("===>>" + str(e))
        self.internetWork = False
        return


      #save local df copy
      self.df = df_final
      #Manage the online results
      manage_kegg_query(self)

      #add extra column with description of kegg name
      ### KEGG KO ###
      if( hasattr(self, 'query_ko') ):
        position_to_insert = df_final.columns.get_loc("KEGG_ko")+1
        #crearte a new empty column
        df_final.insert(loc=position_to_insert, column="KO name", value=['' for i in range(df_final.shape[0])])
        #rename to use
        self.df_ko.rename(columns={0:'Name', 1:'Description'}, inplace=True)
        #mapper
        mapper = self.df_ko.set_index('Name')['Description'].to_dict()
        regex = '|'.join(self.df_ko['Name'])
        # 'ko:123|ko:111|ko:222|ko:333|ko:444|ko:555'
        df_final['KO name'] = df_final['KEGG_ko'].str.replace(regex, lambda m: mapper.get(m.group()), regex=True)

      ### KEGG Pathway ###
      if( hasattr(self, 'query_pathway') ):
        #get position for new column
        position_to_insert = df_final.columns.get_loc("KEGG_Pathway")+1
        #crearte a new empty column
        df_final.insert(loc=position_to_insert, column="Pathway name", value=['' for i in range(df_final.shape[0])])
        #rename to use
        self.df_pathway.rename(columns={0:'Name', 1:'Description'}, inplace=True)
        #mapper
        mapper = self.df_pathway.set_index('Name')['Description'].to_dict()
        regex = '|'.join(self.df_pathway['Name'])
        # 'ko:123|ko:111|ko:222|ko:333|ko:444|ko:555'
        df_final['Pathway name'] = df_final['KEGG_Pathway'].str.replace(regex, lambda m: mapper.get(m.group()), regex=True)

      ### KEGG Module ###
      if( hasattr(self, 'query_module') ):
        #get position for new column
        position_to_insert = df_final.columns.get_loc("KEGG_Module")+1
        #crearte a new empty column
        df_final.insert(loc=position_to_insert, column="Module name", value=['' for i in range(df_final.shape[0])])
        #rename to use
        self.df_module.rename(columns={0:'Name', 1:'Description'}, inplace=True)
        #mapper
        mapper = self.df_module.set_index('Name')['Description'].to_dict()
        regex = '|'.join(self.df_module['Name'])
        # 'ko:123|ko:111|ko:222|ko:333|ko:444|ko:555'
        #df_final['Module name'] = df_final['KEGG_Module'].str.replace(regex, lambda m: mapper.get(m.group()), regex=True)
        df_final['Module name'] = df_final['KEGG_Module'].str.replace(',', '|', regex=False).str.replace(regex, lambda m: mapper.get(m.group()), regex=True)

      ### KEGG Reaction ###
      if( hasattr(self, 'query_reaction') ):
        #get position for new column
        position_to_insert = df_final.columns.get_loc("KEGG_Reaction")+1
        #crearte a new empty column
        df_final.insert(loc=position_to_insert, column="Reaction name", value=['' for i in range(df_final.shape[0])])
        #rename to use
        self.df_reaction.rename(columns={0:'Name', 1:'Description'}, inplace=True)
        #mapper
        mapper = self.df_reaction.set_index('Name')['Description'].to_dict()
        regex = '|'.join(self.df_reaction['Name'])
        # 'ko:123|ko:111|ko:222|ko:333|ko:444|ko:555'
        #df_final['Reaction name'] = df_final['KEGG_Reaction'].str.replace(regex, lambda m: mapper.get(m.group()), regex=True)
        df_final['Reaction name'] = df_final['KEGG_Reaction'].str.replace(';', '|', regex=False).str.replace(regex, lambda m: mapper.get(m.group()), regex=True)

    #add description to COG column if exist
    if 'COG_category' in df_final.columns:
      #get position for new column
      position_to_insert = df_final.columns.get_loc("COG_category")+1
      #crearte a new empty column
      df_final.insert(loc=position_to_insert, column="COG name", value=['' for i in range(df_final.shape[0])])
      # Funzione per ottenere la descrizione in base alle lettere
      def get_description(lettere):
          lettere = lettere.split(';')
          descrizioni_lista = []

          for gruppo in lettere:
              lettere_gruppo = gruppo.split(',')
              #check if fill empty cells in annotation
              if(window.var_chc_unassigned.get() == 1):
                descrizioni_gruppo = [COG_dict.get(lettera, 'unassigned') for lettera in lettere_gruppo]
              else:
                descrizioni_gruppo = [COG_dict.get(lettera, '') for lettera in lettere_gruppo]
              descrizioni_lista.append(','.join(descrizioni_gruppo))

          return ';'.join(descrizioni_lista)

      # Applicare la funzione al DataFrame
      df_final['COG name'] = df_final['COG_category'].apply(get_description)

    #Add table to dict for aggregation windows
    if(MyUtility.workDict['functional_mode'] == 'dynamic'):
      MyUtility.workDict['functional_table'] = [col for col in df_final_annotation.columns if col not in ['query']]
    else:
      MyUtility.workDict['functional_table']      = ["COG_category", "GOs", "EC", "KEGG_ko", "KEGG_Pathway", "KEGG_Module", "KEGG_Reaction", "CAZy"]
      MyUtility.workDict['functional_to_display'] = ["COG_category", "GOs", "EC", "KEGG KO", "KEGG pathway", "KEGG module", "KEGG reaction", "CAZy"]

    
    

    #save edit df in tmp variable
    window.df_tmp = df_final